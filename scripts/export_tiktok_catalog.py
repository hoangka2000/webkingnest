"""Export products from PostgreSQL into TikTok Seller Center batch upload template."""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import openpyxl

from app.database import Product, get_db, parse_json_field
from app.services import (
    THUNG_PACK_PRICES,
    YEN_CHUNG_PACK_PRICES,
    _product_images,
    facebook_catalog_id,
    facebook_purchase_options,
    normalize_product_type,
    supports_yen_chung_packs,
)

BRAND = "kingnest"
DEFAULT_CATEGORY = "Pantry Food/Bird's Nest"
DEFAULT_INVENTORY = 100
DEFAULT_DELIVERY = "Default"
COUNTRY_OF_ORIGIN = "Vietnam"
TIKTOK_NAME_MIN = 25
TIKTOK_NAME_MAX = 255
NAME_SUFFIX = " | Kingnest Yến Sào Khánh Hoà"
DATA_START_ROW = 6

IMAGE_FIELDS = [
    "main_image",
    "image_2",
    "image_3",
    "image_4",
    "image_5",
    "image_6",
    "image_7",
    "image_8",
    "image_9",
]


def tiktok_purchase_options(product: Product) -> list[dict[str, Any]]:
    if supports_yen_chung_packs(product):
        return [
            {
                "id": facebook_catalog_id(product.id, variant_id),
                "variantId": variant_id,
                "label": label,
                "title": product.title,
                "price": YEN_CHUNG_PACK_PRICES[variant_id],
            }
            for variant_id, label in (("10hu", "10 hũ"), ("30hu", "30 hũ"))
        ]

    product_type = normalize_product_type(product.product_type)
    base_price = int(product.price)

    if product.slug == "thung-yen-gia-si":
        return [
            {
                "id": facebook_catalog_id(product.id, variant_id),
                "variantId": variant_id,
                "label": label,
                "title": product.title,
                "price": THUNG_PACK_PRICES[variant_id],
            }
            for variant_id, label in (("40hu", "40 hũ"), ("60hu", "60 hũ"))
        ]

    if product_type in {"yen-tinh-che", "yen-tho"}:
        return [
            {
                "id": facebook_catalog_id(product.id, variant_id),
                "variantId": variant_id,
                "label": label,
                "title": product.title,
                "price": round(base_price * multiplier),
            }
            for variant_id, label, multiplier in (
                ("50g", "50g", 0.5),
                ("100g", "100g", 1),
            )
        ]

    return facebook_purchase_options(product)


def tiktok_product_name(title: str) -> str:
    name = (title or "").strip()
    if len(name) < TIKTOK_NAME_MIN:
        name = f"{name}{NAME_SUFFIX}"
    if len(name) < TIKTOK_NAME_MIN:
        name = f"{name} | Yến Sào Khánh Hoà Nguyên Chất Chính Hãng"
    return name[:TIKTOK_NAME_MAX]


def tiktok_category(product: Product) -> str:
    return DEFAULT_CATEGORY


def tiktok_description(product: Product) -> str:
    parts: list[str] = []

    short_desc = (product.short_desc or "").strip()
    if short_desc:
        parts.append(short_desc)

    benefits = parse_json_field(product.benefits, [])
    if isinstance(benefits, list) and benefits:
        parts.append("Lợi ích:")
        parts.extend(f"• {item}" for item in benefits[:5])

    usage = parse_json_field(product.usage, [])
    if isinstance(usage, list) and usage:
        parts.append("Cách dùng:")
        parts.extend(f"• {item}" for item in usage[:5])

    specs = parse_json_field(product.specs, {})
    if isinstance(specs, dict) and specs:
        parts.append("Thông số:")
        for key, value in list(specs.items())[:6]:
            parts.append(f"• {key}: {value}")

    if not parts:
        parts.append(product.title)

    return "\n".join(parts)[:5000]


def parcel_defaults(product: Product) -> tuple[int, int, int, int]:
    product_type = normalize_product_type(product.product_type)
    slug = product.slug or ""

    if product_type == "hop-qua":
        if "10-hu" in slug:
            return 1200, 30, 25, 12
        return 800, 28, 22, 10

    if product_type == "yen-chung":
        return 500, 25, 20, 15

    if slug == "thung-yen-gia-si":
        return 5000, 45, 35, 25

    if product_type in {"yen-tinh-che", "yen-tho"}:
        return 200, 20, 15, 8

    return 300, 22, 18, 10


def variant_property_name(product: Product, option_count: int) -> str:
    if option_count <= 1:
        return ""

    product_type = normalize_product_type(product.product_type)
    if supports_yen_chung_packs(product) or product.slug == "thung-yen-gia-si":
        return "Quy cách"
    if product_type in {"yen-tinh-che", "yen-tho"}:
        return "Khối lượng"
    return "Quy cách"


def row_values(product: Product, option: dict[str, Any], option_count: int) -> list[Any]:
    _main_image, gallery = _product_images(product)
    images = gallery[: len(IMAGE_FIELDS)]
    while len(images) < len(IMAGE_FIELDS):
        images.append(None)

    weight, length, width, height = parcel_defaults(product)
    property_name = variant_property_name(product, option_count)
    property_value = option.get("label") if property_name else None

    return [
        tiktok_category(product),
        BRAND,
        tiktok_product_name(option.get("title") or product.title),
        tiktok_description(product),
        *images,
        property_name or None,
        property_value,
        images[0],
        None,
        None,
        weight,
        length,
        width,
        height,
        DEFAULT_DELIVERY,
        int(option["price"]),
        DEFAULT_INVENTORY,
        option["id"],
        None,
        None,
        COUNTRY_OF_ORIGIN,
    ]


def build_rows(products: list[Product]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for product in products:
        options = tiktok_purchase_options(product)
        option_count = len(options)
        for option in options:
            rows.append(row_values(product, option, option_count))
    return rows


def export_tiktok_catalog(template_path: Path, output_path: Path) -> int:
    with get_db() as db:
        products = db.query(Product).order_by(Product.id.asc()).all()
        rows = build_rows(products)

    workbook = openpyxl.load_workbook(template_path)
    worksheet = workbook["Template"]

    max_row = worksheet.max_row
    if max_row >= DATA_START_ROW:
        worksheet.delete_rows(DATA_START_ROW, max_row - DATA_START_ROW + 1)

    for index, values in enumerate(rows):
        row_number = DATA_START_ROW + index
        for column_index, value in enumerate(values, start=1):
            worksheet.cell(row=row_number, column=column_index, value=value)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()
    return len(rows)


def main() -> None:
    template = ROOT / "Tiktoksellercenter_batchupload_20260705_template.xlsx"
    output = template

    if len(sys.argv) > 1:
        output = Path(sys.argv[1]).expanduser().resolve()
    if len(sys.argv) > 2:
        template = Path(sys.argv[2]).expanduser().resolve()

    count = export_tiktok_catalog(template, output)
    print(f"Exported {count} rows to {output}")


if __name__ == "__main__":
    main()
